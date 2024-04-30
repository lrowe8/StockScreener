from pathlib import Path
import pandas as pd
import openpyxl
import datetime
import yfinance as yf

if __name__ == '__main__':
    path = Path('Data')
    path.mkdir(parents=True, exist_ok=True)

    # Query Dates
    start = datetime.datetime(2010, 1, 1)
    end = datetime.datetime.now()

    wb = openpyxl.load_workbook('Data.xlsx')
    ws = wb['Current Stocks']

    header = {ws[f'{cell}1'].value : cell for cell in 'ABCDE'}

    for i, row in enumerate(ws.rows, 1):
        pulled_today = True

        if i != 1:
            date = ws[f'{header["Last Pull Date"]}{i}'].value if ws[f'{header["Last Pull Date"]}{i}'].value else datetime.datetime(1970, 1, 1) 
            symbol = ws[f'{header["Symbol"]}{i}'].value
            xlsx_path = path / f'{symbol}.xlsx'

            # Get the previous figures
            if xlsx_path.exists():
                previous_pull = pd.read_excel(xlsx_path, index_col='Date')
            else:
                previous_pull = pd.DataFrame()

            # If today is different from the last pull
            if datetime.datetime.now().date() != date.date():
                # Preserve the previous values
                ws[f'{header["Last Pull Date"]}{i}'] = datetime.datetime.now()
                ws[f'{header["Last 20 Day"]}{i}'] = previous_pull['Close'].rolling(20).mean().iloc[-1] if not previous_pull.empty else 0
                ws[f'{header["Last 50 Day"]}{i}'] = previous_pull['Close'].rolling(20).mean().iloc[-1] if not previous_pull.empty else 0
                ws[f'{header["Last 200 Day"]}{i}'] = previous_pull['Close'].rolling(20).mean().iloc[-1] if not previous_pull.empty else 0

                # Get from yahoo
                result = yf.download(symbol, start, end)
                # Save the file
                result.to_excel(xlsx_path)
            else:
                result = previous_pull

            # Calculate rolling averages
            this_20 = result['Close'].rolling(20).mean()
            this_50 = result['Close'].rolling(50).mean()
            this_200 = result['Close'].rolling(200).mean()

            # The result to graph will be the last 90 days
            result = result[result.index >= (datetime.datetime.now() - datetime.timedelta(days=90))]

            if all([
                this_20.iloc[-1] > ws[f'{header["Last 20 Day"]}{i}'].value, 
                this_50.iloc[-1] > ws[f'{header["Last 50 Day"]}{i}'].value,
                this_200.iloc[-1] > ws[f'{header["Last 200 Day"]}{i}'].value 
            ]):
                color = 'MediumSeaGreen'
            elif all([
                this_20.iloc[-1] < ws[f'{header["Last 20 Day"]}{i}'].value, 
                this_50.iloc[-1] < ws[f'{header["Last 50 Day"]}{i}'].value,
                this_200.iloc[-1] < ws[f'{header["Last 200 Day"]}{i}'].value
            ]):
                color = 'Tomato'
            else:
                color = 'Black'

    wb.save('Data.xlsx')

            








    # First index is for row

    # ws[2][header['Symbol']].value # Get value
    # ws['B2'].value # Get value
    # datetime.date.today() # Date without time
    # ws['B2'].value.date() == datetime.date.today() # Compare dates